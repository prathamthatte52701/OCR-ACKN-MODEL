import asyncio
import os
import re
import threading
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from filelock import FileLock
from openpyxl.worksheet.worksheet import Worksheet

EXPORT_DIR = Path(__file__).resolve().parents[3] / "exports"
HEADERS = ["Document Type", "Number", "Date", "Timestamp"]
MONTHS = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]

_DATE_RE = re.compile(r"^(\d{2})/(\d{2})/(\d{4})$")

# Workbook writes are serialized per physical filename, two-layer:
#   1. An asyncio.Lock keyed by filename - cheap, sub-millisecond, handles the
#      common case (multiple concurrent requests within this one process).
#   2. An OS-level FileLock (filelock package) on a sidecar .lock file -
#      handles the rarer but real case of two backend processes somehow
#      bound to the same port at once (happened during dev testing here),
#      where two independent asyncio.Lock instances in two processes cannot
#      see each other and would otherwise silently interleave read-modify-
#      write cycles, dropping rows. Per-filename (not global) so saves to
#      different users'/years' workbooks never wait on each other.
_locks_guard = threading.Lock()
_write_locks: dict[str, asyncio.Lock] = {}


def _get_write_lock(filename: str) -> asyncio.Lock:
    with _locks_guard:
        lock = _write_locks.get(filename)
        if lock is None:
            lock = asyncio.Lock()
            _write_locks[filename] = lock
        return lock


class FileLockedError(Exception):
    pass


def _ensure_export_dir() -> None:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def file_path(filename: str) -> Path:
    # No path traversal via a user-supplied filename.
    safe = os.path.basename(filename)
    if not safe.endswith(".xlsx"):
        safe = f"{safe}.xlsx"
    return EXPORT_DIR / safe


def current_period(now: datetime | None = None) -> tuple[int, str]:
    now = now or datetime.now()
    return now.year, MONTHS[now.month - 1]


def month_from_date(date_str: str | None, fallback_now: datetime | None = None) -> str:
    """Worksheet month comes from the DOCUMENT'S OWN extracted date, not
    today's date - a document dated 30/06 always lands in the June sheet
    even if saved in July. Falls back to the current month if unparseable."""
    if date_str:
        match = _DATE_RE.match(date_str)
        if match:
            mm = int(match.group(2))
            if 1 <= mm <= 12:
                return MONTHS[mm - 1]
    return current_period(fallback_now)[1]


def _add_header_row(sheet: Worksheet) -> None:
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)


def _lock_path(target: Path) -> Path:
    return target.with_suffix(".lock")


def _create_workbook_sync(filename: str, month: str) -> Path:
    """Assumes the caller already holds this filename's cross-process
    FileLock - does not acquire it itself, so it can be called both
    standalone (via _create_workbook_locked_sync) and inline from inside
    _append_row_sync's own lock without a reentrant/nested acquisition."""
    _ensure_export_dir()
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    sheet = workbook.create_sheet(month)
    _add_header_row(sheet)
    target = file_path(filename)
    tmp_target = target.with_suffix(f".tmp{os.getpid()}.xlsx")
    try:
        workbook.save(tmp_target)
        os.replace(tmp_target, target)
    except PermissionError as exc:
        tmp_target.unlink(missing_ok=True)
        raise FileLockedError(
            f'"{target.name}" is open in Excel. Close it and try saving again.'
        ) from exc
    except Exception:
        tmp_target.unlink(missing_ok=True)
        raise
    return target


def _create_workbook_locked_sync(filename: str, month: str) -> Path:
    _ensure_export_dir()
    target = file_path(filename)
    with FileLock(str(_lock_path(target)), timeout=30):
        return _create_workbook_sync(filename, month)


async def create_workbook(filename: str, month: str | None = None) -> Path:
    month = month or current_period()[1]
    async with _get_write_lock(filename):
        return await asyncio.to_thread(_create_workbook_locked_sync, filename, month)


def _format_number_cell(row: dict) -> str | None:
    if row["documentType"] == "Tax Invoice":
        parts = [p for p in (row.get("taxInvoiceNo"), row.get("referenceNo")) if p]
        return " / ".join(parts) if parts else None
    return row.get("number") or None


def _append_row_sync(filename: str, month: str, row: dict) -> Path:
    _ensure_export_dir()
    target = file_path(filename)

    # Cross-process lock, held for the full read-modify-write below - closes
    # the gap an in-process-only asyncio.Lock leaves if two backend
    # processes ever end up bound to the same port (seen during dev here):
    # two independent asyncio.Lock instances can't see each other and would
    # otherwise let both processes load-modify-save concurrently, each
    # unaware of the other's write, silently dropping whichever row lost the
    # race to os.replace.
    with FileLock(str(_lock_path(target)), timeout=30):
        if not target.exists():
            # Settings can point at a file that was deleted from disk - recreate it.
            _create_workbook_sync(filename, month)

        # Write to a temp file and swap it in with os.replace, rather than
        # openpyxl.save(target) directly - if the target is locked open elsewhere
        # (e.g. Excel), a direct save can truncate the file before the lock error
        # surfaces, corrupting it for every save after the lock is released.
        tmp_target = target.with_suffix(f".tmp{os.getpid()}.xlsx")
        try:
            workbook = openpyxl.load_workbook(target)
            if month in workbook.sheetnames:
                sheet = workbook[month]
            else:
                sheet = workbook.create_sheet(month)
                _add_header_row(sheet)
            sheet.append(
                [row["documentType"], _format_number_cell(row), row["date"], row["timestamp"]]
            )
            workbook.save(tmp_target)
            os.replace(tmp_target, target)
        except PermissionError as exc:
            tmp_target.unlink(missing_ok=True)
            raise FileLockedError(
                f'"{target.name}" is open in Excel. Close it and try saving again.'
            ) from exc
        except Exception:
            tmp_target.unlink(missing_ok=True)
            raise
        return target


async def append_row(filename: str, month: str, row: dict) -> Path:
    # Per-filename asyncio.Lock first (fast path, handles the common case of
    # concurrent requests within this one process without ever touching the
    # filesystem lock); the FileLock inside _append_row_sync is the second,
    # cross-process layer. Keyed by filename (not global) so concurrent
    # saves to different users'/years' workbooks never wait on each other.
    async with _get_write_lock(filename):
        return await asyncio.to_thread(_append_row_sync, filename, month, row)


def _remove_rows_sync(filename: str, rows_to_remove: list[dict]) -> tuple[int, bool]:
    """Surgically deletes specific rows rather than rewriting the whole file.
    Rows have no stable ID (see _append_row_sync), so matching is by the same
    (Document Type, Number, Date) tuple _format_number_cell already produces
    for the Number column - grouped into a per-sheet Counter (multiset, not a
    set) because "Save Again" can legitimately append more than one identical
    row for the same document, and only as many rows as are actually being
    purged should come out, not every row that happens to match. Within a
    sheet, matched rows are deleted bottom-most-index-first so an earlier
    delete_rows() call never shifts the index of a row still pending
    deletion. Returns (rows_removed, workbook_is_now_fully_empty)."""
    target = file_path(filename)
    if not target.exists() or not rows_to_remove:
        return 0, False

    by_sheet: dict[str, Counter] = defaultdict(Counter)
    for row in rows_to_remove:
        sheet_name = month_from_date(row.get("date"))
        key = (row["documentType"], _format_number_cell(row), row.get("date"))
        by_sheet[sheet_name][key] += 1

    removed = 0
    fully_empty = False
    # Cross-process lock, same as _append_row_sync - a concurrent save/remove
    # against this same physical file must never interleave with this
    # read-modify-write.
    with FileLock(str(_lock_path(target)), timeout=30):
        if not target.exists():
            return 0, False
        workbook = openpyxl.load_workbook(target)
        for sheet_name, counter in by_sheet.items():
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            remaining = Counter(counter)
            rows_to_delete: list[int] = []
            for row_cells in sheet.iter_rows(min_row=2):
                key = (row_cells[0].value, row_cells[1].value, row_cells[2].value)
                if remaining.get(key, 0) > 0:
                    rows_to_delete.append(row_cells[0].row)
                    remaining[key] -= 1
            for idx in sorted(rows_to_delete, reverse=True):
                sheet.delete_rows(idx)
            removed += len(rows_to_delete)

        fully_empty = all(sheet.max_row <= 1 for sheet in workbook.worksheets)
        if not fully_empty:
            tmp_target = target.with_suffix(f".tmp{os.getpid()}.xlsx")
            try:
                workbook.save(tmp_target)
                os.replace(tmp_target, target)
            except PermissionError as exc:
                tmp_target.unlink(missing_ok=True)
                raise FileLockedError(
                    f'"{target.name}" is open in Excel. Close it and try saving again.'
                ) from exc
            except Exception:
                tmp_target.unlink(missing_ok=True)
                raise

    # Unlink AFTER the FileLock context exits, not inside it - on Windows a
    # file with an open handle (the lock file itself, held by this same
    # process for the `with` block above) cannot be deleted while that
    # handle is still open.
    if fully_empty:
        target.unlink(missing_ok=True)
        _lock_path(target).unlink(missing_ok=True)

    return removed, fully_empty


async def remove_rows(filename: str, rows_to_remove: list[dict]) -> tuple[int, bool]:
    async with _get_write_lock(filename):
        return await asyncio.to_thread(_remove_rows_sync, filename, rows_to_remove)
